"""
Утилиты для экспорта данных в Excel
"""
import io
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

def export_to_excel(
    data: List[Dict[str, Any]], 
    filename: str = "export.xlsx",
    sheet_name: str = "Data",
    headers: Optional[Dict[str, str]] = None
) -> bytes:
    """
    Экспортирует данные в Excel файл
    
    Args:
        data: Список словарей с данными
        filename: Имя файла
        sheet_name: Имя листа
        headers: Словарь для переименования колонок {old_name: new_name}
    
    Returns:
        bytes: Содержимое Excel файла
    """
    try:
        if not data:
            # Создаем пустой файл если нет данных
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(["Нет данных для экспорта"])
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        
        # Создаем DataFrame из данных
        df = pd.DataFrame(data)
        
        # Переименовываем колонки если указаны headers
        if headers:
            df = df.rename(columns=headers)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Добавляем данные
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel file created successfully: {filename}, rows: {len(data)}")
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {str(e)}")
        # Возвращаем простой файл с ошибкой
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws.append([f"Ошибка при создании файла: {str(e)}"])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

def export_analytics_to_excel(
    analytics_data: Dict[str, Any],
    filename: str = "analytics.xlsx"
) -> bytes:
    """
    Экспортирует аналитические данные в Excel с несколькими листами
    
    Args:
        analytics_data: Словарь с аналитическими данными
        filename: Имя файла
    
    Returns:
        bytes: Содержимое Excel файла
    """
    try:
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Создаем листы для разных типов данных
        sheets_created = False
        
        # Лист с общей статистикой
        if 'summary' in analytics_data:
            ws_summary = wb.create_sheet("Общая статистика")
            summary_data = analytics_data['summary']
            
            # Добавляем заголовки
            ws_summary.append(["Метрика", "Значение"])
            
            # Добавляем данные
            for key, value in summary_data.items():
                ws_summary.append([key, value])
            
            sheets_created = True
        
        # Лист с пользователями
        if 'users' in analytics_data and analytics_data['users']:
            ws_users = wb.create_sheet("Пользователи")
            users_data = analytics_data['users']
            
            if isinstance(users_data, list) and users_data:
                # Добавляем заголовки
                headers = list(users_data[0].keys())
                ws_users.append(headers)
                
                # Добавляем данные
                for user in users_data:
                    row = [user.get(header, '') for header in headers]
                    ws_users.append(row)
                
                sheets_created = True
        
        # Лист с генерациями
        if 'generations' in analytics_data and analytics_data['generations']:
            ws_generations = wb.create_sheet("Генерации")
            generations_data = analytics_data['generations']
            
            if isinstance(generations_data, list) and generations_data:
                # Добавляем заголовки
                headers = list(generations_data[0].keys())
                ws_generations.append(headers)
                
                # Добавляем данные
                for generation in generations_data:
                    row = [generation.get(header, '') for header in headers]
                    ws_generations.append(row)
                
                sheets_created = True
        
        # Если не создали ни одного листа, создаем пустой
        if not sheets_created:
            ws_empty = wb.create_sheet("Нет данных")
            ws_empty.append(["Нет данных для экспорта"])
        
        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Analytics Excel file created successfully: {filename}")
        return buffer.getvalue()
        
    except Exception as e:
        logger.error(f"Error creating analytics Excel file: {str(e)}")
        # Возвращаем простой файл с ошибкой
        wb = Workbook()
        ws = wb.active
        ws.title = "Error"
        ws.append([f"Ошибка при создании файла: {str(e)}"])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
